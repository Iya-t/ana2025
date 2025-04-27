from flask import Flask, request, jsonify, send_from_directory
from flask_cors import CORS
from openpyxl import Workbook, load_workbook
from datetime import datetime
import os

app = Flask(__name__, static_folder=".", static_url_path="")
CORS(app)  # Можно добавить настройки, если нужно ограничить доступ по определённым доменам

EXCEL_FILE = 'responses.xlsx'

def create_excel_if_not_exists():
    """Создаём файл Excel, если он не существует."""
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.append(["Дата и время"] + [f"Вопрос {i}" for i in range(1, 11)] + ["Сумма", "Комментарий"])
        wb.save(EXCEL_FILE)

@app.route('/')
def index():
    """Возвращаем статическую страницу index.html"""
    return send_from_directory('.', 'index.html')

@app.route('/analyze', methods=['POST'])
def analyze():
    """Обработка отправленных данных и запись их в файл Excel."""
    try:
        # Получаем данные из запроса
        data = request.get_json(force=True)
        print("Получены данные:", data)

        # Обработка ответов
        answers = [int(data.get(f'q{i}', 0)) for i in range(1, 11)]
        score = sum(answers)
        comment = "Скорее всего, всё в порядке." if score < 13 else "Возможно, есть признаки депрессии."

        # Открываем и обновляем Excel файл
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        ws.append([now] + answers + [score, comment])
        wb.save(EXCEL_FILE)

        # Отправляем ответ
        return jsonify({"message": f"Ваш балл: {score}. {comment}"})
    except Exception as e:
        print("Ошибка:", e)
        return jsonify({"message": "Ошибка на сервере", "error": str(e)}), 500

if __name__ == '__main__':
    create_excel_if_not_exists()  # Проверяем, существует ли файл
    app.run(host='0.0.0.0', port=int(os.environ.get('PORT', 5000)), debug=True)
